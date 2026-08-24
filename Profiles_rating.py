"""
rate_pairs_openrouter.py

Sends each dating-profile pair to one or more models via OpenRouter, at
whatever temperature the model is actually configured with (no explicit
temperature override — see TEMPERATURE below). Each pair's questionnaire
.docx (Pair_XX_questionnaire.docx — intro, Male Profile, Female Profile,
and all 13 questions, already combined) is read directly and its full
text becomes the ONE single message sent to the model. No system prompt
is used, and no part of the wording is hardcoded in this script —
everything comes straight from the .docx files in QUESTIONNAIRES_DIR.

Every pair is rated once per model in MODELS — so total calls =
len(MODELS) x number of pairs. Each call is completely fresh and
isolated: a brand-new `messages` list with no history carried over from
any other pair or model, and each response is parsed independently.

Results (the 12 numeric scores plus the 5 ranked factors from question
13) are written directly into spreadsheets — one per model, saved in a
results/ folder — no per-pair .json or .txt files are created.

Requirements:
    pip install openai python-dotenv python-docx openpyxl --break-system-packages

Usage:
    Create a .env file in this same directory containing:
        OPENROUTER_API_KEY=sk-or-...
    Then just run:
        python rate_pairs_openrouter.py

    Edit MODELS near the top of this file to control which model(s) run.

Input:
    profiles/Pair_01_questionnaire.docx
    profiles/Pair_02_questionnaire.docx
    ...
    (one fully-assembled questionnaire per pair — nothing else needed)

Output:
    results/<model>_results.xlsx   (one file per model, one row per pair)
"""

import os
import re
import time
from pathlib import Path

import docx
from dotenv import load_dotenv
from openai import OpenAI, APIConnectionError, APITimeoutError, InternalServerError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()  # reads a .env file in the current directory into os.environ, if present

# --- Model selection -------------------------------------------------------
# Any OpenRouter model slug works here, e.g.:
#   "openai/gpt-5.5"
#   "anthropic/claude-opus-4.8"
#   "google/gemini-3.7-flash"
# See https://openrouter.ai/models for the current list/slugs.
# Every pair is rated once per model listed here — edit this list directly
# to add, remove, or swap which models run.
MODELS = [
    "meta/muse-spark-1.2"
]

# Temperature is intentionally NOT set on the API call — this leaves each
# model at whatever temperature it's actually configured with by default
# on OpenRouter. Set this to a float instead if you want one fixed
# temperature applied to every call.
TEMPERATURE = None

# Maximum length of the model's response. If a model writes a lot of
# explanatory reasoning (especially for low-compatibility pairs, where
# models tend to over-explain), a low limit here can cut the response off
# before it reaches later questions — those then show up as blank in the
# spreadsheet even though nothing went wrong with parsing. This is the
# STARTING budget for every call; see MAX_TOKENS_CEILING below for what
# happens if even this isn't enough for a particular pair.
MAX_TOKENS = 4000

# If a response comes back truncated (the model hit MAX_TOKENS before
# finishing), that specific call is automatically retried with a bigger
# budget — doubling each time — rather than requiring you to guess the
# right fixed number for every possible pair up front. This is the hard
# ceiling on that doubling: 4000 -> 8000 -> 16000, then stop. Raise this
# if you have pairs that are still truncating even at 16000 tokens
# (unusual for this questionnaire, but possible with a very verbose model).
MAX_TOKENS_CEILING = 16000

# Transient network/connection errors are retried automatically (things
# like a dropped SSL connection, a timeout, or the provider being briefly
# unavailable) — this many attempts per (pair, model) call before giving
# up and moving on, so one flaky connection doesn't crash the whole run.
MAX_RETRIES = 4
RETRY_BACKOFF_SECONDS = 5  # doubles after each retry: 5s, 10s, 20s, ...

QUESTIONNAIRES_DIR = Path("profiles")
RESULTS_DIR = Path("results")
RESULTS_DIR.mkdir(exist_ok=True)

# The title heading each generated questionnaire starts with, e.g.
# "Profile Pair Compatibility Questionnaire — Pair 1". This is skipped
# when building the prompt — it's a filing label for humans, not part of
# the actual task text sent to the model.
TITLE_HEADING_STYLE = "Heading 1"


def extract_prompt_text(path):
    """
    Reads a Pair_XX_questionnaire.docx top to bottom and returns its full
    body text (intro, Male Profile, Female Profile, all 13 questions) as
    a single string, exactly as written in the document — skipping only
    the title heading line. Nothing about the wording is duplicated or
    hardcoded in this script.
    """
    d = docx.Document(path)
    lines = []
    for p in d.paragraphs:
        text = p.text.strip()
        if not text:
            continue
        style_name = p.style.name if p.style else ""
        if style_name == TITLE_HEADING_STYLE:
            continue
        lines.append(text)
    return "\n\n".join(lines)


def discover_questionnaires():
    """Find every Pair_XX_questionnaire.docx in QUESTIONNAIRES_DIR (profiles/), in order."""
    files = sorted(QUESTIONNAIRES_DIR.glob("Pair_*_questionnaire.docx"))
    if not files:
        raise SystemExit(
            f"No questionnaire .docx files found in {QUESTIONNAIRES_DIR}/ "
            "— check the folder name and that the files are there."
        )
    return files


def rate_one_pair(client, prompt_text, model):
    """
    Makes a single, self-contained chat-completion call for one pair, at
    one model. A brand-new `messages` list — containing exactly one user
    message — is built here every time. Nothing is carried over from
    previous calls (across pairs OR models), and no system message is
    used at all.

    Two independent failure modes are handled automatically:

    1. Transient network errors (dropped connections, timeouts, brief
       provider outages) are retried up to MAX_RETRIES times with
       increasing delay between attempts. Other errors (bad API key,
       invalid model name, etc.) are NOT retried — they're raised
       immediately since retrying won't fix them.

    2. Truncated responses (the model hit the token limit before
       finishing) are retried with a DOUBLED token budget, up to
       MAX_TOKENS_CEILING. This means you don't have to guess the right
       fixed MAX_TOKENS for every pair — a normally-short pair still
       uses the small starting budget, while a pair that genuinely needs
       more room gets it automatically, on that same pair's call only.
    """
    current_max_tokens = MAX_TOKENS

    while True:
        kwargs = dict(
            model=model,
            messages=[
                {"role": "user", "content": prompt_text}
            ],
            max_tokens=current_max_tokens,
        )
        if TEMPERATURE is not None:
            kwargs["temperature"] = TEMPERATURE

        last_error = None
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                response = client.chat.completions.create(**kwargs)
                choice = response.choices[0]
                truncated = choice.finish_reason == "length"

                if truncated and current_max_tokens < MAX_TOKENS_CEILING:
                    next_max_tokens = min(current_max_tokens * 2, MAX_TOKENS_CEILING)
                    print(f"  Truncated at max_tokens={current_max_tokens}, "
                          f"retrying this pair with max_tokens={next_max_tokens} ...")
                    current_max_tokens = next_max_tokens
                    break  # exit the network-retry loop, re-enter the while loop with a bigger budget

                if truncated:
                    # Already at the ceiling and still truncated — give up
                    # escalating and return what we have, clearly flagged.
                    print(f"  WARNING: still truncated at MAX_TOKENS_CEILING="
                          f"{MAX_TOKENS_CEILING}. Later questions/factors for "
                          f"this pair will be blank. Raise MAX_TOKENS_CEILING "
                          f"if this happens often.")

                return choice.message.content, truncated

            except (APIConnectionError, APITimeoutError, InternalServerError) as e:
                last_error = e
                if attempt < MAX_RETRIES:
                    wait = RETRY_BACKOFF_SECONDS * (2 ** (attempt - 1))
                    print(f"  Network error ({type(e).__name__}), "
                          f"retrying in {wait}s (attempt {attempt}/{MAX_RETRIES}) ...")
                    time.sleep(wait)
                else:
                    print(f"  Giving up after {MAX_RETRIES} attempts: {e}")
                    raise last_error


def parse_numeric_fields(raw_text):
    """
    Parses the model's numbered answers into a flat dict.

    This works in two structural steps rather than one flat set of
    regexes, which matters because models vary in how they label each
    item — some write plain "7. 35", others write "**Item 7 (Hobbies
    compatibility):** 35". A naive "search for '7.' anywhere in the
    text" approach is unsafe: question 13's own factor list is numbered
    1-6, so an unbounded search for "7." (etc.) can accidentally match
    inside that list instead of the real item 7 — producing a WRONG
    value, not just a missing one, which is worse.

    Step 1: locate every item marker (in either format) and its
    position, keep only the first occurrence of each item number, and
    use each marker's position as the boundary for where that item's
    value must be found — never past the next marker. This makes the
    parse immune to numbers that appear later in the response (like
    question 13's list) leaking backward into earlier items.

    Step 2: item 13's marker position is used as the split point between
    "the 12 scored items" and "the factors list" — so factor items 1-6
    are parsed completely separately.
    """
    fields = {}

    # Matches either "Item N" (optionally bold, optionally followed by a
    # parenthetical description and/or colon) or a plain "N." at the
    # start of a line. Both forms are supported because different models
    # (and even the same model across pairs) phrase this differently.
    item_marker_re = re.compile(
        r"(?:\*{0,2}\s*Item\s+(?P<n1>\d{1,2})\b(?:\s*\([^)]{0,100}\))?\s*:?\s*\*{0,2}"
        r"|(?:^|\n)\s*\**\s*(?P<n2>\d{1,2})\.(?!\d))",
        re.IGNORECASE
    )

    raw_markers = []
    for m in item_marker_re.finditer(raw_text):
        n = m.group("n1") or m.group("n2")
        if n is not None:
            raw_markers.append((int(n), m.start(), m.end()))

    # Keep only the first occurrence of each item number, in the order
    # they appear — this is what prevents e.g. question 13's internal
    # "1. ..." factor list from being mistaken for item 1's own marker,
    # since the real item 1 marker (whichever format) always appears
    # earlier in the response.
    seen = set()
    markers = []
    for num, start, end in raw_markers:
        if num not in seen:
            seen.add(num)
            markers.append((num, start, end))

    marker_by_num = {num: (start, end) for num, start, end in markers}

    def value_window(item_num):
        """
        The text this item's value must be found in: from just after
        this item's own marker, up to the start of whichever marker
        comes next in the response (of any item number) — never
        further. Returns "" if this item's marker wasn't found at all.
        """
        if item_num not in marker_by_num:
            return ""
        _, label_end = marker_by_num[item_num]
        later_starts = [start for num, start, _ in markers if start > label_end]
        window_end = min(later_starts) if later_starts else len(raw_text)
        return raw_text[label_end:window_end]

    def extract_value(item_num):
        window = value_window(item_num)
        if not window:
            return None
        if item_num == 1:
            m = re.search(r"\b([1-7])\b", window)
        elif item_num == 2:
            m = re.search(r"\b([1-3]|[a-cA-C])\b", window)
        else:
            m = re.search(r"\b(\d{1,3})\b", window)
        return m.group(1) if m else None

    field_names = {
        1: "q1_realism", 2: "q2_construction", 3: "q3_social_demo",
        4: "q4_social_demo_conf", 5: "q5_psych", 6: "q6_psych_conf",
        7: "q7_hobbies", 8: "q8_hobbies_conf", 9: "q9_expectations",
        10: "q10_expectations_conf", 11: "q11_overall", 12: "q12_overall_conf",
    }
    for num, name in field_names.items():
        fields[name] = extract_value(num)

    # Question 13 is free text: up to 6 ranked factors (a 6th "Other"
    # bucket is allowed), each with a direction (increased/decreased)
    # and a points value that should sum to 100 across all factors.
    # Its marker position (found above) is used as the split point, so
    # this never overlaps with the items-1-12 parsing above.
    if 13 in marker_by_num:
        _, q13_start = marker_by_num[13]
        q13_block = raw_text[q13_start:].strip()
    else:
        # Fallback for the rare case the marker-finder didn't catch a
        # non-standard "13" label — same simple approach as before.
        q13_match = re.search(r"13\.(.*)", raw_text, re.DOTALL)
        q13_block = q13_match.group(1).strip() if q13_match else ""

    # Models often add a trailing summary after the numbered list (a
    # "Total: 100 points" line, a "---" divider, a closing note). None of
    # that belongs to the last factor, so it's cut off here before the
    # list is split into individual factors — otherwise it all gets
    # glued onto factor 5/6's text.
    trailer_match = re.search(
        r"\n\s*(?:\*{0,2}Total\b|-{3,}|\*Summary\b|Summary note)",
        q13_block, re.IGNORECASE
    )
    if trailer_match:
        q13_block = q13_block[:trailer_match.start()]

    # Numbered items are matched directly so any preamble text before the
    # first numbered item (like "Top five factors:") is discarded rather
    # than counted as factor #1.
    factor_matches = re.findall(
        r"\d+[\.\)]\s*(.+?)(?=\n\s*\d+[\.\)]|\Z)", q13_block, re.DOTALL
    )
    if not factor_matches:
        factor_matches = re.findall(r"[-•]\s*(.+?)(?=\n\s*[-•]|\Z)", q13_block, re.DOTALL)
    factor_lines = [f.strip().replace("\n", " ") for f in factor_matches if f.strip()]

    def extract_points(factor_text):
        """
        Pulls the points value out of a factor's text, trying a few
        common ways a model might phrase it before falling back to the
        last standalone number in the text. Returns None if nothing
        number-like is found at all.
        """
        m = re.search(r"(\d{1,3})\s*(?:points?|pts?)\b", factor_text, re.IGNORECASE)
        if m:
            return m.group(1)
        m = re.search(r"\((\d{1,3})\)\s*$", factor_text)
        if m:
            return m.group(1)
        m = re.search(r"(\d{1,3})(?!.*\d)", factor_text)
        return m.group(1) if m else None

    def clean_factor_description(factor_text):
        """
        Strips the points phrase (e.g. "— 35 points", "(35 pts)",
        "Points: 35") and markdown bold markers out of the factor text,
        since the number already lives in its own column — this keeps
        the description column pure prose instead of duplicating the
        points value inline.
        """
        # Remove a "<dash/em-dash> NN points" chunk
        text = re.sub(
            r"\s*[-—–]\s*\d{1,3}\s*(?:points?|pts?)\b",
            "", factor_text, flags=re.IGNORECASE
        )
        # Remove a "(NN points)" parenthetical
        text = re.sub(
            r"\s*\(\d{1,3}\s*(?:points?|pts?)?\)\s*",
            " ", text, flags=re.IGNORECASE
        )
        # Remove a trailing "Points: NN" or "Point value: NN" label
        # (allowing for markdown bold markers on either side, e.g.
        # "**Points: 38**")
        text = re.sub(
            r"\s*\**\s*Points?(?:\s+value)?\s*:\s*\d{1,3}\s*\**\s*$",
            "", text, flags=re.IGNORECASE
        )
        # Strip markdown bold/italic markers
        text = text.replace("**", "").replace("*", "")
        return text.strip(" -—–\u2014\u2013")

    for i in range(6):  # up to 5 factors + optional 6th "Other" bucket
        if i < len(factor_lines):
            fields[f"q13_factor_{i+1}"] = clean_factor_description(factor_lines[i])
            fields[f"q13_factor_{i+1}_points"] = extract_points(factor_lines[i])
        else:
            fields[f"q13_factor_{i+1}"] = None
            fields[f"q13_factor_{i+1}_points"] = None

    fields["raw_response"] = raw_text
    return fields


# Column layout for the output spreadsheet: (header label, field key)
# No "Model" column — each spreadsheet is already scoped to one model.
# Each Q13 factor gets its own column, immediately followed by a column
# for that factor's points value, so the two sit side by side and are
# easy to scan (and to sum/chart) separately from the free-text factor
# description itself.
COLUMNS = [
    ("Pair", "pair_id"),
    ("Truncated?", "truncated"),
    ("Incomplete?", "incomplete"),
    ("Q1 Realism (1-7)", "q1_realism"),
    ("Q2 Construction (1-3)", "q2_construction"),
    ("Q3 Social/Demo Compat", "q3_social_demo"),
    ("Q4 Social/Demo Conf.", "q4_social_demo_conf"),
    ("Q5 Psych Compat", "q5_psych"),
    ("Q6 Psych Conf.", "q6_psych_conf"),
    ("Q7 Hobbies Compat", "q7_hobbies"),
    ("Q8 Hobbies Conf.", "q8_hobbies_conf"),
    ("Q9 Expectations Compat", "q9_expectations"),
    ("Q10 Expectations Conf.", "q10_expectations_conf"),
    ("Q11 Overall Compat", "q11_overall"),
    ("Q12 Overall Conf.", "q12_overall_conf"),
    ("Q13 Factor 1", "q13_factor_1"),
    ("Q13 Factor 1 Points", "q13_factor_1_points"),
    ("Q13 Factor 2", "q13_factor_2"),
    ("Q13 Factor 2 Points", "q13_factor_2_points"),
    ("Q13 Factor 3", "q13_factor_3"),
    ("Q13 Factor 3 Points", "q13_factor_3_points"),
    ("Q13 Factor 4", "q13_factor_4"),
    ("Q13 Factor 4 Points", "q13_factor_4_points"),
    ("Q13 Factor 5", "q13_factor_5"),
    ("Q13 Factor 5 Points", "q13_factor_5_points"),
    ("Q13 Factor 6 (Other)", "q13_factor_6"),
    ("Q13 Factor 6 Points", "q13_factor_6_points"),
    ("Raw Response", "raw_response"),
]


def model_filename(model):
    """
    Turns an OpenRouter model slug into a filesystem-safe filename, e.g.
    'openai/gpt-5.5' -> 'openai_gpt-5.5_results.xlsx'.
    """
    safe = model.replace("/", "_")
    return f"{safe}_results.xlsx"


def build_workbook(rows):
    """
    Builds a formatted spreadsheet from the collected rows — one row per
    (pair, model) combination, all scores and factors as columns.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    FONT = "Arial"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    cell_font = Font(name=FONT, size=10)
    pair_font = Font(name=FONT, bold=True, size=10)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    for c, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30

    for r, row in enumerate(rows, start=2):
        for c, (_, key) in enumerate(COLUMNS, start=1):
            val = row.get(key)
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            if key == "pair_id":
                cell.font = pair_font
                cell.alignment = center
            elif key == "raw_response":
                cell.font = cell_font
                cell.alignment = wrap
            elif key.startswith("q13_factor") and key.endswith("_points"):
                cell.font = cell_font
                cell.alignment = center
            elif key.startswith("q13_factor"):
                cell.font = cell_font
                cell.alignment = wrap
            else:
                cell.font = cell_font
                cell.alignment = center

    ws.column_dimensions["A"].width = 10
    for c, (_, key) in enumerate(COLUMNS, start=1):
        if key == "pair_id":
            continue
        col_letter = get_column_letter(c)
        if key == "raw_response":
            ws.column_dimensions[col_letter].width = 80
        elif key.startswith("q13_factor") and key.endswith("_points"):
            ws.column_dimensions[col_letter].width = 12
        elif key.startswith("q13_factor"):
            ws.column_dimensions[col_letter].width = 32
        elif key in ("truncated", "incomplete"):
            ws.column_dimensions[col_letter].width = 11
        else:
            ws.column_dimensions[col_letter].width = 13

    ws.freeze_panes = "D2"

    # Print setup: fit to page width, landscape, repeat header row
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:1"

    return wb


def main():
    api_key = os.environ.get("OPENROUTER_API_KEY")
    if not api_key:
        raise SystemExit("Set OPENROUTER_API_KEY (e.g. in a .env file) before running.")

    client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
        default_headers={
            "HTTP-Referer": "https://your-site-or-project.example",  # replace or remove
            "X-Title": "Profile Pair Compatibility Rating",
        },
    )

    questionnaire_files = discover_questionnaires()
    print(f"Found {len(questionnaire_files)} questionnaires: "
          f"{[f.stem for f in questionnaire_files]}")

    print(f"Running models: {MODELS}")
    total_calls = len(MODELS) * len(questionnaire_files)
    print(f"Total API calls this run: {total_calls}")

    for model in MODELS:
        print(f"\n=== Model: {model} ===")

        rows = []
        out_path = RESULTS_DIR / model_filename(model)

        for qfile in questionnaire_files:
            pair_id = qfile.stem.replace("_questionnaire", "")  # e.g. "Pair_01"

            prompt_text = extract_prompt_text(qfile)

            print(f"Rating {pair_id} ({model}) ...")
            try:
                # Fresh, isolated call every time: new pair, new model, new
                # messages list — nothing carried over from any other call.
                raw_text, truncated = rate_one_pair(client, prompt_text, model=model)
                parsed = parse_numeric_fields(raw_text)

                # Distinguish two different failure modes that can both
                # leave later columns blank:
                #   - "Truncated?" = the model was cut off (finish_reason
                #     == "length") — a token-budget problem, already
                #     retried with more room by rate_one_pair.
                #   - "Incomplete?" = the model finished normally
                #     (finish_reason == "stop") but never answered later
                #     questions anyway — NOT a token-budget problem, so
                #     raising MAX_TOKENS won't fix this one. Check the Raw
                #     Response column to see what the model actually said.
                key_late_fields = ["q7_hobbies", "q9_expectations", "q11_overall", "q13_factor_1"]
                incomplete = (not truncated) and all(parsed.get(k) is None for k in key_late_fields)

                row = {
                    "pair_id": pair_id,
                    "truncated": "YES" if truncated else "",
                    "incomplete": "YES" if incomplete else "",
                    **parsed,
                }
            except Exception as e:
                # After MAX_RETRIES failed attempts (or a non-retryable
                # error), don't crash the whole run — record the failure
                # for this pair and move on to the next one. You can spot
                # these afterward: every score column will be blank and
                # q13_factor_1 will contain the error message.
                print(f"  FAILED for {pair_id} after retries: {e}")
                row = {"pair_id": pair_id, "q13_factor_1": f"ERROR: {e}"}

            rows.append(row)

            # Save after every pair, not just at the end of the model —
            # so if the run is interrupted, everything done so far for
            # this model is already on disk.
            wb = build_workbook(rows)
            wb.save(out_path)

            time.sleep(1)  # gentle pacing; adjust/remove based on your rate limits

        print(f"Saved {len(rows)} rows to {out_path}")

    print(f"\nDone. One spreadsheet per model saved under {RESULTS_DIR}/")


if __name__ == "__main__":
    main()